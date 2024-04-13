# For this project, we are updating all of the values in one column of an excel spreadsheet.
# We will use openpyxl to open the excel file, then some of the chart methods to build a chart

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Putting all of the code within a function so that we can simply pass a file name and have it updated

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    # Selecting the sheet
    sheet = wb['Sheet1']

    # these are methods for reading an individual sheet
    # cell = sheet['a1']
    # cell = sheet.cell(1,1)
    # check how many rows in the sheet
    # print(sheet.max_row)

    # Create a title for the column with the new prices, then iterate through the rows to get each original price
    # and update it to the corrected price, then store the corrected price in column 4

    sheet.cell(1, 4).value = "corrected_price"
    for row in range(2,sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Gathering the values to build the bar chart

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    # Building the bar chart and telling the program where to put it on the excel sheet, note the cell listed is the
    # where the top left corner of the chart will be placed

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    # Save the file, in this situation we are saving it under the original file name because we want to overwrite
    # the files

    wb.save(filename)